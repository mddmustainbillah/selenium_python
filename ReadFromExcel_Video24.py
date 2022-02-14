import openpyxl

path = "C:/Users/Titan Tech-2/PycharmProjects/SeleniumWithPython/New folder/salary.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active  # workbook.get_sheet_by_name("Sheet1") //
# if excel file has more than one sheet then we have to specify the sheet name Code will be
# """workbook.get_sheet_by_name("Sheet1")""" instead of """workbook.active"""

# C:/Users/Titan Tech-2/PycharmProjects/SeleniumWithPython/salary.xlsx
rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for row in range(1, rows+1):
    for col in range(1, cols+1):
        print(sheet.cell(row=row, column=col).value, end="        ")
    print()